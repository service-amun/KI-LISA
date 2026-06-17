# © 2026 Karola Fromm-Nasreldin | AILIZA — Alle Rechte vorbehalten
"""
AILIZA — Datei-Upload & Text-Extraktion
Unterstützt: PDF, Word (.docx), Excel (.xlsx/.xls), CSV, TXT, Bilder (JPG/PNG)
"""

import io
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse

router = APIRouter()

MAX_DATEIGRÖSSE = 10 * 1024 * 1024  # 10 MB


def _pdf_text(daten: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(daten))
        seiten = [p.extract_text() or "" for p in reader.pages]
        text = "\n\n".join(s.strip() for s in seiten if s.strip())
        return text[:15000] or "(PDF enthält keinen lesbaren Text — möglicherweise gescannt)"
    except Exception as e:
        raise HTTPException(400, f"PDF konnte nicht gelesen werden: {e}")


def _word_text(daten: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(daten))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text[:15000] or "(Word-Dokument ist leer)"
    except Exception as e:
        raise HTTPException(400, f"Word-Datei konnte nicht gelesen werden: {e}")


def _excel_text(daten: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(daten), read_only=True, data_only=True)
        teile = []
        for blatt in wb.sheetnames[:5]:  # max 5 Tabellenblätter
            ws = wb[blatt]
            zeilen = []
            for zeile in ws.iter_rows(max_row=200, values_only=True):
                zellen = [str(z) if z is not None else "" for z in zeile]
                if any(z.strip() for z in zellen):
                    zeilen.append(" | ".join(zellen))
            if zeilen:
                teile.append(f"=== Tabellenblatt: {blatt} ===\n" + "\n".join(zeilen))
        return "\n\n".join(teile)[:15000] or "(Excel-Datei ist leer)"
    except Exception as e:
        raise HTTPException(400, f"Excel-Datei konnte nicht gelesen werden: {e}")


def _bild_info(daten: bytes, dateiname: str) -> str:
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(daten))
        return (
            f"[Bild hochgeladen: {dateiname}]\n"
            f"Größe: {img.width}×{img.height} Pixel, Format: {img.format}\n\n"
            "Hinweis: Automatische Texterkennung (OCR) ist für Bilder noch nicht aktiv. "
            "Beschreibe kurz was auf dem Bild zu sehen ist, oder stelle deine Frage dazu."
        )
    except Exception:
        return f"[Bild hochgeladen: {dateiname}] — Bild konnte nicht analysiert werden."


@router.post("/upload")
async def datei_hochladen(datei: UploadFile = File(...)):
    """
    Datei hochladen und Text extrahieren.
    Unterstützt: PDF, Word, Excel, CSV, TXT, JPG, PNG
    """
    if datei.size and datei.size > MAX_DATEIGRÖSSE:
        raise HTTPException(413, "Datei zu groß (max. 10 MB)")

    name = datei.filename or "unbekannt"
    endung = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    daten = await datei.read()

    if endung == "pdf":
        text = _pdf_text(daten)
        typ = "pdf"
    elif endung in ("docx", "doc"):
        text = _word_text(daten)
        typ = "word"
    elif endung in ("xlsx", "xls"):
        text = _excel_text(daten)
        typ = "excel"
    elif endung == "csv":
        text = daten.decode("utf-8", errors="replace")[:15000]
        typ = "csv"
    elif endung == "txt":
        text = daten.decode("utf-8", errors="replace")[:15000]
        typ = "text"
    elif endung in ("jpg", "jpeg", "png", "webp", "bmp"):
        text = _bild_info(daten, name)
        typ = "bild"
    else:
        raise HTTPException(415, f"Dateityp '.{endung}' wird nicht unterstützt. Erlaubt: PDF, Word, Excel, CSV, TXT, JPG, PNG")

    return {
        "dateiname": name,
        "typ": typ,
        "zeichen": len(text),
        "vorschau": text[:300] + ("…" if len(text) > 300 else ""),
        "text": text,
    }
